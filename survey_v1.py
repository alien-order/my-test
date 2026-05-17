#!/usr/bin/env python3
"""
survey_v1.py — PROD 단일 스키마 Oracle 테이블 현황 조사

출력 컬럼: 번호, 시스템명, 테이블명, 테이블설명, 사용여부, 데이터보관최초일자,
           1개월평균데이터생성건수, 칼럼수(개), 전체테이블용량
"""

import os
import re
import sys
import shutil
from datetime import datetime
from typing import Dict, List, Optional, Tuple

try:
    import oracledb
except ImportError:
    print("[ERROR] oracledb 패키지 없음: pip install oracledb")
    sys.exit(1)

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter, column_index_from_string
    from dateutil import parser as dateutil_parser
except ImportError:
    print("[ERROR] 패키지 없음: pip install pandas openpyxl python-dateutil")
    sys.exit(1)


# =============================================================================
# ★ 설정
# =============================================================================
ORACLE_CLIENT_LIB = r"C:\oracle\instantclient_21_9"

DB_CONFIGS = {
    "prod": {
        "host":         "PROD_HOST",
        "port":         1521,
        "service_name": "PROD_SERVICE",
        "user":         "PROD_USER",
        "password":     "PROD_PASSWORD",
    }
}

SYSTEM_NAME   = "TEST"
SAMPLE_ROWS   = 10

# 마스터 테이블 패턴
MASTER_PATTERN = (
    "CODE|COCD|_CD$|_CD_|MST|MASTER|BASE|STD|CMM|COMMON|"
    "TYPE|KIND|STAT|STATUS|GRP|GROUP|CLASS|CATG|CONFIG|PARAM|META|"
    "MENU|ROLE|AUTH|PERMISSION|POLICY"
)

# VARCHAR 날짜 패턴 (최소 길이, regex, strptime 포맷)
VARCHAR_DATE_PATTERNS = [
    (8,  r"^\d{8}$",               "%Y%m%d"),
    (12, r"^\d{12}$",              "%Y%m%d%H%M"),
    (14, r"^\d{14}$",              "%Y%m%d%H%M%S"),
    (16, r"^\d{16}$",              "%Y%m%d%H%M%S%f"),  # 16자리 → microsec 앞 2자리
    (17, r"^\d{17,}$",             "%Y%m%d%H%M%S%f"),
    (10, r"^\d{4}-\d{2}-\d{2}$",  "%Y-%m-%d"),
    (19, r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$", "%Y-%m-%d %H:%M:%S"),
]

LINK_FONT = Font(color="0563C1", underline="single", bold=True)


# =============================================================================
# 유틸
# =============================================================================
def fmt_gb(bytes_val) -> str:
    if not bytes_val:
        return "0.0GB"
    gb = bytes_val / (1024 ** 3)
    if gb <= 0:
        return "0.0GB"
    if gb < 0.1:
        return "0.1GB"
    return f"{gb:.1f}GB"


def make_conn(cfg: Dict):
    dsn = f"{cfg['host']}:{cfg['port']}/{cfg['service_name']}"
    return oracledb.connect(user=cfg["user"], password=cfg["password"], dsn=dsn)


def fetchall_dict(cur) -> List[Dict]:
    cols = [d[0].lower() for d in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]


def parse_cell(ref: str) -> Tuple[int, int]:
    """'B4' → (col=2, row=4)"""
    m = re.match(r"^([A-Za-z]+)(\d+)$", ref.strip())
    if not m:
        raise ValueError(f"잘못된 셀 참조: {ref}")
    return column_index_from_string(m.group(1)), int(m.group(2))


# =============================================================================
# 사용여부 판단 (Q3 + ASH + DML + 마스터패턴, 8순위)
# =============================================================================
USAGE_SQL = """
WITH
q3_select AS (
    SELECT  sp.object_name                AS table_name,
            MAX(sa.last_active_time)      AS last_select_time
    FROM    v$sql_plan sp
    JOIN    v$sqlarea  sa ON sa.sql_id = sp.sql_id
    WHERE   sp.object_owner = USER
      AND   sp.object_name  IS NOT NULL
      AND   sp.object_type  LIKE 'TABLE%%'
    GROUP   BY sp.object_name
),
ash_access AS (
    SELECT  o.object_name                                              AS table_name,
            COUNT(CASE WHEN ash.sql_opname = 'SELECT' THEN 1 END)     AS ash_select,
            COUNT(CASE WHEN ash.sql_opname IN (
                       'INSERT','UPDATE','DELETE','MERGE') THEN 1 END) AS ash_dml
    FROM    dba_hist_active_sess_history ash
    JOIN    dba_objects o
            ON  o.object_id  = ash.current_obj#
            AND o.owner       = USER
            AND o.object_type = 'TABLE'
    WHERE   ash.sample_time >= SYSDATE - 8
      AND   ash.current_obj# > 0
    GROUP   BY o.object_name
),
dml_hist AS (
    SELECT  table_name,
            SUM(inserts) AS inserts,
            SUM(updates) AS updates,
            SUM(deletes) AS deletes
    FROM    user_tab_modifications
    WHERE   partition_name IS NULL
    GROUP   BY table_name
),
master_chk AS (
    SELECT  table_name,
            CASE WHEN REGEXP_LIKE(table_name,
                 'CODE|COCD|_CD$|_CD_|MST|MASTER|BASE|STD|CMM|COMMON|'||
                 'TYPE|KIND|STAT|STATUS|GRP|GROUP|CLASS|CATG|CONFIG|PARAM|META|'||
                 'MENU|ROLE|AUTH|PERMISSION|POLICY', 'i')
                 THEN 'Y' ELSE 'N' END AS is_master
    FROM    user_tables
)
SELECT
    t.table_name,
    NVL(t.num_rows, -1)                                              AS num_rows,
    CASE
        WHEN ash.ash_select > 0
            THEN 'Y (ASH-SELECT확인)'
        WHEN ash.ash_dml > 0
            THEN 'Y (ASH-DML확인)'
        WHEN q3.table_name IS NOT NULL
            THEN 'Y (Q3-SELECT흔적)'
        WHEN NVL(m.inserts,0)+NVL(m.updates,0)+NVL(m.deletes,0) > 0
            THEN 'Y (DML이력)'
        WHEN t.num_rows = 0
            THEN 'N'
        WHEN t.num_rows IS NULL
            THEN '확인필요 (통계미수집)'
        WHEN t.num_rows > 0 AND mc.is_master = 'Y'
            THEN '확인필요 (마스터추정)'
        WHEN t.num_rows > 0
            THEN '확인필요'
        ELSE 'N'
    END AS usage_flag
FROM    user_tables t
LEFT JOIN dml_hist  m   ON  m.table_name  = t.table_name
LEFT JOIN q3_select q3  ON  q3.table_name = t.table_name
LEFT JOIN ash_access ash ON ash.table_name = t.table_name
LEFT JOIN master_chk mc  ON  mc.table_name = t.table_name
"""


def load_usage_map(conn) -> Dict[str, Dict]:
    """테이블명 → {usage_flag, num_rows} 딕셔너리 반환"""
    cur = conn.cursor()
    try:
        # FLUSH: 최신 DML 이력 반영
        try:
            cur.execute("BEGIN DBMS_STATS.FLUSH_DATABASE_MONITORING_INFO; END;")
        except Exception:
            pass
        cur.execute(USAGE_SQL)
        rows = fetchall_dict(cur)
        return {r["table_name"]: r for r in rows}
    except Exception as e:
        print(f"  [WARN] 사용여부 쿼리 실패 ({e}) — 단순 DML 이력만 사용")
        return {}
    finally:
        cur.close()


# =============================================================================
# 테이블 목록
# =============================================================================
def get_table_list(conn) -> List[Dict]:
    cur = conn.cursor()
    cur.execute("""
        SELECT t.table_name,
               NVL(tc.comments, '-') AS comments,
               t.num_rows,
               o.created             AS created_dt
        FROM   user_tables      t
        LEFT JOIN user_tab_comments tc ON tc.table_name = t.table_name
        LEFT JOIN user_objects      o  ON  o.object_name = t.table_name
                                       AND o.object_type  = 'TABLE'
        ORDER BY t.table_name
    """)
    return fetchall_dict(cur)


# =============================================================================
# 컬럼 수
# =============================================================================
def get_column_count(conn, table_name: str) -> int:
    cur = conn.cursor()
    cur.execute(
        "SELECT COUNT(*) FROM user_tab_columns WHERE table_name = :1",
        [table_name]
    )
    row = cur.fetchone()
    cur.close()
    return row[0] if row else 0


# =============================================================================
# 용량 (테이블 + LOB 세그먼트 합산)
# =============================================================================
def get_size_bytes(conn, table_name: str) -> int:
    cur = conn.cursor()
    cur.execute("""
        SELECT NVL(SUM(bytes), 0)
        FROM (
            SELECT bytes FROM user_segments WHERE segment_name = :1
            UNION ALL
            SELECT s.bytes
            FROM   user_lobs l
            JOIN   user_segments s ON s.segment_name = l.segment_name
            WHERE  l.table_name = :2
        )
    """, [table_name, table_name])
    row = cur.fetchone()
    cur.close()
    return row[0] if row else 0


# =============================================================================
# 1개월 평균 INSERT 건수
# =============================================================================
def get_monthly_avg_insert(conn, table_name: str) -> str:
    cur = conn.cursor()
    cur.execute("""
        SELECT NVL(SUM(inserts), 0), MIN(timestamp), MAX(timestamp)
        FROM   user_tab_modifications
        WHERE  table_name = :1 AND partition_name IS NULL
    """, [table_name])
    row = cur.fetchone()
    cur.close()
    if not row or row[0] == 0:
        return "0"
    total_ins = row[0]
    min_dt = row[1]
    max_dt = row[2]
    if min_dt and max_dt and min_dt != max_dt:
        months = max(1, round((max_dt - min_dt).days / 30))
    else:
        months = 1
    avg = round(total_ins / months)
    if avg >= 10000:
        return f"{avg:,}"
    return str(avg)


# =============================================================================
# 날짜 컬럼 탐지
# =============================================================================
def _try_parse_varchar_date(val: str) -> bool:
    s = str(val).strip()
    for min_len, pattern, fmt in VARCHAR_DATE_PATTERNS:
        if len(s) >= min_len and re.match(pattern, s):
            try:
                datetime.strptime(s[:len(fmt.replace('%Y','0000').replace('%m','00').replace('%d','00').replace('%H','00').replace('%M','00').replace('%S','00').replace('%f','000000'))], fmt)
                return True
            except Exception:
                pass
    # dateutil fallback
    try:
        dateutil_parser.parse(s, yearfirst=True)
        return True
    except Exception:
        return False


def _is_varchar_date_col(conn, table_name: str, col_name: str) -> bool:
    cur = conn.cursor()
    try:
        cur.execute(
            f"SELECT {col_name} FROM {table_name} WHERE ROWNUM <= 100"
        )
        rows = cur.fetchall()
        if not rows:
            return False
        vals = [r[0] for r in rows if r[0] is not None]
        if not vals:
            return False
        hits = sum(1 for v in vals if _try_parse_varchar_date(str(v)))
        return hits / len(vals) >= 0.8
    except Exception:
        return False
    finally:
        cur.close()


CREATE_PATTERNS = re.compile(
    r"CRE|REG|INS|FIRST|OPEN|START|BEGIN|OCCUR|INIT|ENTR",
    re.IGNORECASE,
)


def detect_date_column(conn, table_name: str) -> Optional[str]:
    """
    우선순위:
    1. DATE/TIMESTAMP 타입 + 생성일 패턴 컬럼명
    2. DATE/TIMESTAMP 타입 (패턴 없음)
    3. VARCHAR이지만 샘플이 날짜 형태
    반환: 컬럼명 or None
    """
    cur = conn.cursor()
    cur.execute("""
        SELECT column_name, data_type
        FROM   user_tab_columns
        WHERE  table_name = :1
        ORDER  BY column_id
    """, [table_name])
    cols = cur.fetchall()
    cur.close()

    date_cols      = []
    date_pat_cols  = []
    varchar_cols   = []

    for col_name, data_type in cols:
        if data_type in ("DATE",) or data_type.startswith("TIMESTAMP"):
            if CREATE_PATTERNS.search(col_name):
                date_pat_cols.append(col_name)
            else:
                date_cols.append(col_name)
        elif "CHAR" in data_type:
            varchar_cols.append(col_name)

    if date_pat_cols:
        return date_pat_cols[0]
    if date_cols:
        return date_cols[0]

    # VARCHAR 날짜 패턴 탐지
    for col_name in varchar_cols:
        if _is_varchar_date_col(conn, table_name, col_name):
            return col_name

    return None


# =============================================================================
# 데이터보관최초일자
# =============================================================================
def get_first_date(conn, table_name: str, num_rows, date_col: Optional[str],
                   created_dt) -> str:
    if num_rows == 0:
        return "-"

    if date_col:
        cur = conn.cursor()
        try:
            cur.execute(
                f"SELECT MIN({date_col}) FROM {table_name}"
            )
            row = cur.fetchone()
            if row and row[0] is not None:
                val = row[0]
                if isinstance(val, datetime):
                    return val.strftime("%Y-%m-%d")
                s = str(val).strip()
                # 8자리 이상 숫자면 날짜 형식으로 변환
                if re.match(r"^\d{8}", s):
                    try:
                        return datetime.strptime(s[:8], "%Y%m%d").strftime("%Y-%m-%d")
                    except Exception:
                        pass
                try:
                    return dateutil_parser.parse(s).strftime("%Y-%m-%d")
                except Exception:
                    return s[:10]
        except Exception:
            pass
        finally:
            cur.close()

    # 날짜 컬럼 없음 → 테이블 생성일
    if created_dt:
        if isinstance(created_dt, datetime):
            return created_dt.strftime("%Y-%m-%d")
        return str(created_dt)[:10]
    return "-"


# =============================================================================
# 샘플 데이터 시트 작성
# =============================================================================
HEADER_FILL = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=9)
ALT_FILL    = PatternFill(start_color="F2F7FF", end_color="F2F7FF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
META_FILL = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
META_FONT = Font(italic=True, color="555555", size=9)
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _read_lob(val):
    if val is None:
        return None
    if hasattr(val, "read"):
        try:
            return val.read()
        except Exception:
            return str(val)
    return val


def fetch_sample(conn, table_name: str, date_col: Optional[str]):
    if date_col:
        sql = (
            f"SELECT * FROM ("
            f"SELECT * FROM {table_name} ORDER BY {date_col} DESC"
            f") WHERE ROWNUM <= {SAMPLE_ROWS}"
        )
        is_ordered = True
    else:
        sql        = f"SELECT * FROM {table_name} WHERE ROWNUM <= {SAMPLE_ROWS}"
        is_ordered = False

    cur = conn.cursor()
    try:
        cur.execute(sql)
        cols  = [d[0] for d in cur.description]
        rows  = cur.fetchall()
        clean = [tuple(_read_lob(v) for v in row) for row in rows]
        return pd.DataFrame(clean, columns=cols), is_ordered
    except Exception as e:
        return pd.DataFrame({"[ERROR]": [str(e)]}), False
    finally:
        cur.close()


def write_sample_sheet(wb, sheet_name: str, df: pd.DataFrame, meta: Dict):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    is_ordered = meta.get("is_ordered", False)
    date_col   = meta.get("date_col", "-")

    sort_note = (
        f"ORDER BY {date_col} DESC → 최근 {SAMPLE_ROWS}건"
        if is_ordered
        else f"※ 날짜 컬럼 없음 — 정렬 보장 없는 임의 {SAMPLE_ROWS}건"
    )
    info = "  |  ".join([
        f"테이블: {meta.get('table_name', '')}",
        f"설명: {meta.get('comment', '-')}",
        f"기준컬럼: {date_col}",
        sort_note,
        f"조회: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    ])
    ws.append([info])
    meta_cell        = ws.cell(row=1, column=1)
    meta_cell.fill   = META_FILL if is_ordered else WARN_FILL
    meta_cell.font   = META_FONT
    n_cols = max(len(df.columns), 5) if not df.empty else 5
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    if df.empty:
        ws.append(["(데이터 없음)"])
        return
    if "[ERROR]" in df.columns:
        ws.append([str(df.iloc[0, 0])])
        return

    ws.append(list(df.columns))
    for cell in ws[2]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER

    for r_idx, row in enumerate(df.itertuples(index=False), start=3):
        ws.append(list(row))
        for cell in ws[r_idx]:
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if r_idx % 2 == 0:
                cell.fill = ALT_FILL

    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, 50)

    ws.freeze_panes = "A3"


# =============================================================================
# 메인
# =============================================================================
def main():
    # ── Oracle 클라이언트 초기화 ──
    try:
        oracledb.init_oracle_client(lib_dir=ORACLE_CLIENT_LIB)
        print(f"[Oracle Client] {ORACLE_CLIENT_LIB}")
    except Exception as e:
        print(f"[ERROR] Oracle Client 초기화 실패: {e}")
        sys.exit(1)

    # ── 입력 ──
    template_path = input("기존 Excel 파일 경로: ").strip().strip('"')
    start_cell    = input("번호 시작 셀 (기본 B4): ").strip() or "B4"

    if not os.path.isfile(template_path):
        print(f"[ERROR] 파일 없음: {template_path}")
        sys.exit(1)

    start_col, start_row = parse_cell(start_cell)

    # ── DB 접속 ──
    try:
        conn = make_conn(DB_CONFIGS["prod"])
        print(f"[PROD] 접속 성공")
    except Exception as e:
        print(f"[ERROR] DB 접속 실패: {e}")
        sys.exit(1)

    # ── 사용여부 맵 일괄 조회 ──
    print("[사용여부] 쿼리 실행 중...")
    usage_map = load_usage_map(conn)

    # ── 테이블 목록 ──
    tables = get_table_list(conn)
    print(f"[테이블] {len(tables)}개 발견")

    # ── 기존 엑셀 로드 ──
    wb         = load_workbook(template_path)
    ws         = wb.active
    data_sheet = ws.title

    records = []

    for idx, tbl in enumerate(tables, start=1):
        tname      = tbl["table_name"]
        comment    = tbl["comments"] or "-"
        num_rows   = tbl["num_rows"]
        created_dt = tbl["created_dt"]

        print(f"  [{idx:4d}/{len(tables)}] {tname}", end="", flush=True)

        usage_info = usage_map.get(tname, {})
        usage_flag = usage_info.get("usage_flag", "확인필요")
        if num_rows is None:
            db_num_rows = None
        else:
            db_num_rows = int(num_rows)

        date_col   = detect_date_column(conn, tname)
        first_date = get_first_date(conn, tname, db_num_rows if db_num_rows is not None else -1,
                                    date_col, created_dt)
        monthly    = get_monthly_avg_insert(conn, tname)
        col_count  = get_column_count(conn, tname)
        size_bytes = get_size_bytes(conn, tname)
        size_str   = fmt_gb(size_bytes)

        print(f"  사용여부={usage_flag[:4]}  최초일={first_date}  용량={size_str}")

        records.append({
            "번호":               idx,
            "시스템명":           SYSTEM_NAME,
            "테이블명":           tname,
            "테이블설명":         comment,
            "사용여부":           usage_flag,
            "데이터보관최초일자": first_date,
            "1개월평균데이터생성건수": monthly,
            "칼럼수(개)":         col_count,
            "전체테이블용량":     size_str,
            "_date_col":          date_col,
        })

    # ── 엑셀 데이터 기록 ──
    COLS = [
        "번호", "시스템명", "테이블명", "테이블설명", "사용여부",
        "데이터보관최초일자", "1개월평균데이터생성건수", "칼럼수(개)", "전체테이블용량",
    ]

    for row_offset, rec in enumerate(records):
        excel_row = start_row + row_offset
        for col_offset, col_key in enumerate(COLS):
            cell = ws.cell(row=excel_row, column=start_col + col_offset)
            cell.value = rec[col_key]

    # ── 샘플 시트 생성 ──
    print("\n[샘플 시트] 생성 중...")
    for rec in records:
        tname    = rec["테이블명"]
        date_col = rec["_date_col"]
        df, is_ordered = fetch_sample(conn, tname, date_col)
        write_sample_sheet(wb, str(rec["번호"]), df, {
            "table_name":  tname,
            "comment":     rec["테이블설명"],
            "date_col":    date_col or "-",
            "is_ordered":  is_ordered,
        })

    conn.close()

    # ── 번호 셀에 하이퍼링크 ──
    for row_offset, rec in enumerate(records):
        excel_row  = start_row + row_offset
        sheet_name = str(rec["번호"])
        cell           = ws.cell(row=excel_row, column=start_col)
        cell.value     = rec["번호"]
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font      = LINK_FONT

    # ── 백업 후 저장 ──
    backup = template_path.replace(".xlsx", f"_backup_{datetime.now().strftime('%H%M%S')}.xlsx")
    shutil.copy2(template_path, backup)
    print(f"[백업] {backup}")

    wb.save(template_path)
    print(f"[저장] {template_path}")
    print(f"완료: {len(records)}개 테이블, 샘플 시트 {len(records)}개")


if __name__ == "__main__":
    main()
