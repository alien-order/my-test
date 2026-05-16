#!/usr/bin/env python3
"""
테이블 샘플 데이터 추출 스크립트

table_survey.py 가 만든 Excel 파일(_prod / _prod+arch)을 읽어
각 테이블의 최근 20건 샘플 데이터를 '번호' 이름의 새 시트로 추가하고,
데이터 시트의 '번호' 셀에 해당 샘플 시트로 이동하는 하이퍼링크를 건다.

사용법:
  python table_sample.py table_survey_20260517_prod.xlsx
  python table_sample.py table_survey_20260517_prod+arch.xlsx

설정:
  HYPERLINK_START_CELL  번호 컬럼의 첫 데이터 셀 (기본: "A2")
                        헤더가 1행이고 A열이 번호 컬럼인 경우 A2 가 기본값.
                        예: 제목행이 하나 더 있으면 "A3" 으로 변경.

날짜 컬럼이 있는 테이블: ORDER BY 날짜컬럼 DESC → 진짜 최근 N건
날짜 컬럼이 없는 테이블: 정렬 보장 없는 임의 N건 (시트 상단에 안내 표시)

_prod+arch 파일의 경우:
  소재=아카이브만 테이블 → 아카이브DB 에서 샘플 조회
  그 외 테이블          → 운영DB 에서 샘플 조회
"""

import sys
import os
import re
import shutil
from datetime import datetime
from typing import Dict, List, Optional, Tuple

try:
    import oracledb
except ImportError:
    print("[ERROR] oracledb 패키지가 없습니다: pip install oracledb")
    sys.exit(1)

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("[ERROR] pandas/openpyxl 패키지가 없습니다: pip install pandas openpyxl")
    sys.exit(1)


# =============================================================================
# ★ 설정 (table_survey.py 와 동일하게 수정)
# =============================================================================
ORACLE_CLIENT_LIB = r"C:\oracle\instantclient_21_9"

PROD_DB: Dict = {
    "host":         "YOUR_PROD_HOST",
    "port":         1521,
    "service_name": "YOUR_PROD_SERVICE",
    "user":         "YOUR_USER",
    "password":     "YOUR_PASSWORD",
    "label":        "운영DB",
}

ARCHIVE_DB: Dict = {
    "host":         "YOUR_ARCHIVE_HOST",
    "port":         1521,
    "service_name": "YOUR_ARCHIVE_SERVICE",
    "user":         "YOUR_USER",
    "password":     "YOUR_PASSWORD",
    "label":        "아카이브DB",
}

# 번호 컬럼의 첫 데이터 셀 (헤더=1행, 번호=A열 기준)
HYPERLINK_START_CELL = "A2"

SAMPLE_ROWS = 20


# =============================================================================
# 스타일
# =============================================================================
HEADER_FILL = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=9)
ALT_FILL    = PatternFill(start_color="F2F7FF", end_color="F2F7FF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)
META_FILL   = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
META_FONT   = Font(italic=True, color="555555", size=9)
WARN_FILL   = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LINK_FONT   = Font(color="0563C1", underline="single", bold=True)


# =============================================================================
# Oracle 접속
# =============================================================================
def _make_conn(cfg: Dict):
    dsn = f"{cfg['host']}:{cfg['port']}/{cfg['service_name']}"
    return oracledb.connect(user=cfg["user"], password=cfg["password"], dsn=dsn)


def _read_lob(val):
    if val is None:
        return None
    if hasattr(val, "read"):
        try:
            return val.read()
        except Exception:
            return str(val)
    return val


def fetch_sample(conn, table_name: str, date_col: Optional[str]) -> Tuple[pd.DataFrame, bool]:
    """
    (df, is_ordered) 반환.
    date_col 있으면 ORDER BY date_col DESC 후 ROWNUM 적용 → 최근 N건.
    date_col 없으면 ROWNUM 만 적용 → 임의 N건 (정렬 보장 없음).
    """
    if date_col:
        sql = f"""
            SELECT * FROM (
                SELECT * FROM {table_name}
                ORDER  BY {date_col} DESC
            ) WHERE ROWNUM <= {SAMPLE_ROWS}
        """
        is_ordered = True
    else:
        sql        = f"SELECT * FROM {table_name} WHERE ROWNUM <= {SAMPLE_ROWS}"
        is_ordered = False

    cur = conn.cursor()
    try:
        cur.execute(sql)
        cols  = [d[0] for d in cur.description]
        rows  = cur.fetchall()
        clean = [tuple(_read_lob(v) for v in row) for row in rows]
        return pd.DataFrame(clean, columns=cols), is_ordered
    except Exception as e:
        return pd.DataFrame({"[ERROR]": [str(e)]}), False
    finally:
        cur.close()


# =============================================================================
# Excel 시트 작성
# =============================================================================
def write_sample_sheet(
    wb,
    sheet_name: str,
    df: pd.DataFrame,
    meta: Dict,
):
    """wb 에 sheet_name 시트를 생성(또는 덮어쓰기)하고 샘플 데이터를 기록"""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    is_ordered = meta.get("is_ordered", False)
    date_col   = meta.get("date_col", "-")

    sort_note = (
        f"ORDER BY {date_col} DESC → 최근 {SAMPLE_ROWS}건"
        if is_ordered
        else f"※ 날짜 컬럼 없음 — 정렬 보장 없는 임의 {SAMPLE_ROWS}건"
    )

    # ── 1행: 메타 정보 ──
    info = "  |  ".join([
        f"테이블: {meta.get('table_name', '')}",
        f"설명: {meta.get('comment', '-')}",
        f"기준컬럼: {date_col}",
        f"탐지방법: {meta.get('date_method', '-')}",
        sort_note,
        f"조회: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    ])
    ws.append([info])
    meta_cell        = ws.cell(row=1, column=1)
    meta_cell.fill   = META_FILL if is_ordered else WARN_FILL
    meta_cell.font   = META_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df.columns), 5))

    has_error = not df.empty and "[ERROR]" in df.columns

    if df.empty:
        ws.append(["(데이터 없음)"])
        return
    if has_error:
        ws.append([str(df.iloc[0, 0])])
        return

    # ── 2행: 컬럼 헤더 ──
    ws.append(list(df.columns))
    for cell in ws[2]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER

    # ── 3행~: 데이터 ──
    for r_idx, row in enumerate(df.itertuples(index=False), start=3):
        ws.append(list(row))
        for cell in ws[r_idx]:
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if r_idx % 2 == 0:
                cell.fill = ALT_FILL

    # ── 열 너비 자동 조정 ──
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, 50)

    ws.freeze_panes = "A3"


# =============================================================================
# 하이퍼링크 추가
# =============================================================================
def _parse_cell_ref(cell_ref: str) -> Tuple[int, int]:
    """'A2' → (col_index=1, row_index=2)"""
    m = re.match(r"^([A-Za-z]+)(\d+)$", cell_ref.strip())
    if not m:
        raise ValueError(f"잘못된 셀 참조: {cell_ref}")
    return column_index_from_string(m.group(1)), int(m.group(2))


def add_hyperlinks(wb, data_sheet_name: str, number_col: int, start_row: int, numbers: List[int]):
    """
    데이터 시트의 '번호' 컬럼 셀에 각 샘플 시트(str(번호))로 이동하는 하이퍼링크 추가.
    number_col : 번호 컬럼의 열 인덱스 (1-based)
    start_row  : 첫 번째 데이터 행 번호
    numbers    : 번호 목록 (순서대로 start_row, start_row+1, ... 에 대응)
    """
    ws = wb[data_sheet_name]
    for offset, num in enumerate(numbers):
        excel_row  = start_row + offset
        sheet_name = str(num)
        if sheet_name not in wb.sheetnames:
            continue
        cell           = ws.cell(row=excel_row, column=number_col)
        cell.value     = num
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.font      = LINK_FONT


# =============================================================================
# 메인
# =============================================================================
def _clean(val) -> Optional[str]:
    s = str(val).strip()
    return None if s in ("nan", "-", "", "None") else s


def main():
    # ── 설정에서 start cell 오버라이드 허용 ──
    start_cell = HYPERLINK_START_CELL

    # ── Oracle 클라이언트 초기화 ──
    try:
        oracledb.init_oracle_client(lib_dir=ORACLE_CLIENT_LIB)
        print(f"[Oracle Client] Thick 모드: {ORACLE_CLIENT_LIB}")
    except Exception as e:
        print(f"[ERROR] Oracle Client 초기화 실패: {e}")
        sys.exit(1)

    # ── 입력 파일 ──
    if len(sys.argv) >= 2:
        excel_path = sys.argv[1]
    else:
        excel_path = input("survey Excel 파일 경로: ").strip().strip('"')
    if len(sys.argv) >= 3:
        start_cell = sys.argv[2]   # ex) A3

    if not os.path.isfile(excel_path):
        print(f"[ERROR] 파일 없음: {excel_path}")
        sys.exit(1)

    is_combined = "_prod+arch" in os.path.basename(excel_path)
    print(f"\n[파일] {excel_path}")
    print(f"[모드] {'운영+아카이브 합산' if is_combined else '운영DB 전용'}")
    print(f"[번호 시작 셀] {start_cell}")

    # ── 데이터 시트 읽기 ──
    xls         = pd.ExcelFile(excel_path)
    sheet_name  = xls.sheet_names[0]   # 첫 번째 시트 사용
    df_survey   = xls.parse(sheet_name)
    print(f"[시트] '{sheet_name}'  ({len(df_survey)}개 테이블)")

    if "테이블명" not in df_survey.columns:
        print("[ERROR] '테이블명' 컬럼이 없습니다.")
        sys.exit(1)

    has_date_col    = "기준컬럼" in df_survey.columns
    has_date_method = "탐지방법" in df_survey.columns
    has_source      = "소재" in df_survey.columns
    has_number      = "번호" in df_survey.columns

    # ── DB 접속 (필요한 것만) ──
    conn_prod = conn_arch = None
    try:
        conn_prod = _make_conn(PROD_DB)
        print(f"[운영DB] 접속 성공: {PROD_DB['host']}/{PROD_DB['service_name']}")
    except Exception as e:
        print(f"[ERROR] 운영DB 접속 실패: {e}")
        sys.exit(1)

    if is_combined:
        try:
            conn_arch = _make_conn(ARCHIVE_DB)
            print(f"[아카이브DB] 접속 성공: {ARCHIVE_DB['host']}/{ARCHIVE_DB['service_name']}")
        except Exception as e:
            print(f"[WARN] 아카이브DB 접속 실패 — 아카이브 전용 테이블 샘플 불가: {e}")

    # ── 시트 생성 ──
    wb    = load_workbook(excel_path)
    total = len(df_survey)
    written_numbers: List[int] = []

    try:
        for i, row in enumerate(df_survey.itertuples(index=False), start=1):
            tname = _clean(getattr(row, "테이블명", None))
            if not tname:
                continue

            num         = int(getattr(row, "번호")) if has_number else i
            date_col    = _clean(getattr(row, "기준컬럼",  None)) if has_date_col    else None
            date_method = _clean(getattr(row, "탐지방법",  None)) if has_date_method else None
            comment     = _clean(getattr(row, "테이블설명", None)) or "-"
            source      = _clean(getattr(row, "소재",      None)) if has_source      else None

            # 어느 DB 에서 샘플을 뽑을지 결정
            if source == "아카이브만" and conn_arch:
                conn = conn_arch
                db_label = "아카이브DB"
            else:
                conn = conn_prod
                db_label = "운영DB"

            print(f"  [{i:4d}/{total}] {tname}  ({db_label})", end="", flush=True)

            df_sample, is_ordered = fetch_sample(conn, tname, date_col)

            sheet_title = str(num)   # 시트명 = 번호
            write_sample_sheet(
                wb,
                sheet_title,
                df_sample,
                meta={
                    "table_name":  tname,
                    "comment":     comment,
                    "date_col":    date_col or "-",
                    "date_method": date_method or "-",
                    "is_ordered":  is_ordered,
                },
            )
            written_numbers.append(num)

            order_tag = f"ORDER BY {date_col} DESC" if is_ordered else "정렬없음"
            print(f"  →  {len(df_sample)}건  ({order_tag})")

    finally:
        if conn_prod:
            conn_prod.close()
        if conn_arch:
            conn_arch.close()

    # ── 번호 셀에 하이퍼링크 추가 ──
    if has_number and written_numbers:
        try:
            num_col, num_row = _parse_cell_ref(start_cell)
            add_hyperlinks(wb, sheet_name, num_col, num_row, written_numbers)
            print(f"\n[하이퍼링크] '{sheet_name}' 시트 {start_cell} 부터 {len(written_numbers)}개 적용")
        except Exception as e:
            print(f"\n[WARN] 하이퍼링크 적용 실패: {e}")

    # ── 저장 ──
    backup = excel_path.replace(".xlsx", f"_backup_{datetime.now().strftime('%H%M%S')}.xlsx")
    shutil.copy2(excel_path, backup)
    print(f"[백업] {backup}")

    wb.save(excel_path)
    print(f"[저장] {excel_path}")
    print(f"\n샘플 시트: {len(written_numbers)}개 추가됨  (시트명: 번호 숫자)")


if __name__ == "__main__":
    main()
