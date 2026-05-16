#!/usr/bin/env python3
"""
Oracle 테이블 현황 조사 스크립트

출력 파일 (2개):
  table_survey_YYYYMMDD_prod.xlsx       - 운영DB 테이블 현황
  table_survey_YYYYMMDD_prod+arch.xlsx  - 운영+아카이브 합산 현황
    └ 아카이브에만 있는 테이블: 사용여부=N, 소재=아카이브만

출력 컬럼: 번호, 시스템명, 테이블명, 테이블설명, 사용여부,
           데이터보관최초일자, 1개월평균데이터생성건수, 칼럼수(개), 전체테이블용량
"""

import sys
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

try:
    import oracledb
except ImportError:
    print("[ERROR] oracledb 패키지가 없습니다: pip install oracledb")
    sys.exit(1)

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] pandas/openpyxl 패키지가 없습니다: pip install pandas openpyxl")
    sys.exit(1)


# =============================================================================
# ★ 설정 (환경에 맞게 수정)
# =============================================================================
ORACLE_CLIENT_LIB = r"C:\oracle\instantclient_21_9"

SYSTEM_NAME = "TEST"   # 시스템명 고정값

DB_CONFIGS: Dict[str, Dict] = {
    "prod": {
        "host":         "YOUR_PROD_HOST",
        "port":         1521,
        "service_name": "YOUR_PROD_SERVICE",
        "user":         "YOUR_USER",
        "password":     "YOUR_PASSWORD",
        "label":        "운영DB",
    },
    "archive": {
        "host":         "YOUR_ARCHIVE_HOST",
        "port":         1521,
        "service_name": "YOUR_ARCHIVE_SERVICE",
        "user":         "YOUR_USER",
        "password":     "YOUR_PASSWORD",
        "label":        "아카이브DB",
    },
}

RECENT_MONTHS    = 3
SAMPLE_SIZE      = 100
DETECT_THRESHOLD = 0.80

_DATE_SUFFIX     = datetime.now().strftime("%Y%m%d")
OUTPUT_PROD      = f"table_survey_{_DATE_SUFFIX}_prod.xlsx"
OUTPUT_COMBINED  = f"table_survey_{_DATE_SUFFIX}_prod+arch.xlsx"


# =============================================================================
# 내부 상수
# =============================================================================
VARCHAR_DATE_PATTERNS: List[Tuple[str, str, int, str, str]] = [
    (r"^\d{16,17}$", "%Y%m%d%H%M%S", 14, "YYYYMMDDHHMMSS+", "YYYYMMDDHH24MISS"),
    (r"^\d{14}$",    "%Y%m%d%H%M%S", 14, "YYYYMMDDHHMMSS",  "YYYYMMDDHH24MISS"),
    (r"^\d{12}$",    "%Y%m%d%H%M",   12, "YYYYMMDDHHMM",    "YYYYMMDDHH24MI"),
    (r"^\d{8}$",     "%Y%m%d",        8, "YYYYMMDD",         "YYYYMMDD"),
    (r"^\d{6}$",     "%Y%m",          6, "YYYYMM",           "YYYYMM"),
    (r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$",
     "%Y-%m-%d %H:%M:%S", 19, "YYYY-MM-DD HH:MI:SS", "YYYY-MM-DD HH24:MI:SS"),
    (r"^\d{4}-\d{2}-\d{2}$", "%Y-%m-%d", 10, "YYYY-MM-DD", "YYYY-MM-DD"),
]

DATE_NAME_WEIGHTS: List[Tuple[str, int]] = [
    ("CRE",   10), ("CRTN",  10),
    ("REG",    9),
    ("INS",    8),
    ("FIRST",  7), ("OPEN",   6),
    ("_DT",    5), ("DT_",    5), ("_DATE", 4), ("DATE_", 4),
    ("TIME",   3),
    ("_YMD",   2), ("YMD",    2),
    ("_YM",    1), ("YM_",    1),
]

DATE_TYPES = {
    "DATE", "TIMESTAMP", "TIMESTAMP(0)", "TIMESTAMP(1)", "TIMESTAMP(2)",
    "TIMESTAMP(3)", "TIMESTAMP(4)", "TIMESTAMP(5)", "TIMESTAMP(6)",
    "TIMESTAMP WITH TIME ZONE", "TIMESTAMP WITH LOCAL TIME ZONE",
}


# =============================================================================
# 유틸리티
# =============================================================================
def _name_weight(col_name: str) -> int:
    upper = col_name.upper()
    for pattern, weight in DATE_NAME_WEIGHTS:
        if pattern in upper:
            return weight
    return 0


def _try_parse_date(val: str, fmt: str, slice_len: int) -> bool:
    try:
        dt = datetime.strptime(str(val)[:slice_len], fmt)
        return 1900 <= dt.year <= 2100
    except (ValueError, TypeError):
        return False


def fmt_gb(bytes_val: Optional[float]) -> str:
    """
    바이트 → GB 문자열.
    0 이면 0.0GB, 완전한 0이 아니면 최소 0.1GB로 표기.
    """
    if not bytes_val:
        return "0.0GB"
    gb = bytes_val / (1024 ** 3)
    if gb <= 0:
        return "0.0GB"
    if gb < 0.1:
        return "0.1GB"   # 완전한 0이 아니면 최소 0.1GB
    return f"{gb:.1f}GB"


def fmt_date(val) -> str:
    if val is None:
        return "-"
    if isinstance(val, str):
        return val[:10] if val else "-"
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val)[:10]


# =============================================================================
# OracleInspector
# =============================================================================
class OracleInspector:
    def __init__(self, db_key: str):
        self.db_key = db_key
        self.cfg    = DB_CONFIGS[db_key]
        self.label  = self.cfg["label"]
        self.conn   = None

    def connect(self):
        dsn = f"{self.cfg['host']}:{self.cfg['port']}/{self.cfg['service_name']}"
        self.conn = oracledb.connect(
            user=self.cfg["user"], password=self.cfg["password"], dsn=dsn
        )
        print(f"  [{self.label}] 접속 성공: {dsn}")

    def close(self):
        if self.conn:
            self.conn.close()

    def _execute(self, sql: str, params=None) -> List[Any]:
        cur = self.conn.cursor()
        try:
            cur.execute(sql, params or [])
            return cur.fetchall()
        except Exception as e:
            raise RuntimeError(f"SQL 오류: {e}\nSQL: {sql[:200]}") from e
        finally:
            cur.close()

    def _execute_one(self, sql: str, params=None) -> Optional[Any]:
        rows = self._execute(sql, params)
        return rows[0] if rows else None

    def get_table_list(self) -> List[Dict]:
        sql = """
            SELECT
                t.table_name,
                NVL(tc.comments, '-') AS comments,
                col.cnt               AS col_count,
                NVL(seg.bytes, 0)     AS seg_bytes
            FROM
                user_tables t
                LEFT JOIN user_tab_comments tc
                    ON tc.table_name = t.table_name
                LEFT JOIN (
                    SELECT table_name, COUNT(*) AS cnt
                    FROM   user_tab_columns
                    GROUP  BY table_name
                ) col ON col.table_name = t.table_name
                LEFT JOIN (
                    SELECT table_name, SUM(bytes) AS bytes
                    FROM (
                        SELECT segment_name AS table_name, SUM(bytes) AS bytes
                        FROM   user_segments
                        WHERE  segment_type IN (
                                   'TABLE','TABLE PARTITION','TABLE SUBPARTITION')
                        GROUP  BY segment_name
                        UNION ALL
                        SELECT l.table_name, SUM(s.bytes) AS bytes
                        FROM   user_lobs l
                        JOIN   user_segments s
                            ON  s.segment_name = l.segment_name
                            AND s.segment_type IN ('LOBSEGMENT','LOB PARTITION')
                        GROUP  BY l.table_name
                    )
                    GROUP BY table_name
                ) seg ON seg.table_name = t.table_name
            ORDER BY t.table_name
        """
        rows = self._execute(sql)
        return [
            {"table_name": r[0], "comment": r[1], "col_count": r[2], "seg_bytes": r[3]}
            for r in rows
        ]

    def get_columns(self, table_name: str) -> List[Dict]:
        sql = """
            SELECT column_name, data_type,
                   NVL(data_length, 0) AS data_length, column_id
            FROM   user_tab_columns
            WHERE  table_name = :1
            ORDER  BY column_id
        """
        rows = self._execute(sql, [table_name])
        return [{"name": r[0], "type": r[1], "length": r[2], "id": r[3]} for r in rows]

    def _detect_varchar_date_pattern(
        self, table_name: str, col_name: str
    ) -> Optional[Tuple[str, str, str]]:
        try:
            sql = f"""
                SELECT {col_name}
                FROM (
                    SELECT {col_name}
                    FROM   {table_name}
                    WHERE  {col_name} IS NOT NULL
                      AND  REGEXP_LIKE({col_name}, '^[0-9]{{6,}}$|^[0-9]{{4}}-[0-9]{{2}}')
                )
                WHERE ROWNUM <= {SAMPLE_SIZE}
            """
            rows = self._execute(sql)
        except Exception:
            return None

        vals = [str(r[0]).strip() for r in rows if r[0] is not None]
        if not vals:
            return None

        for pattern_re, py_fmt, slice_len, label, ora_fmt in VARCHAR_DATE_PATTERNS:
            matched = sum(
                1 for v in vals
                if re.match(pattern_re, v) and _try_parse_date(v, py_fmt, slice_len)
            )
            if matched / len(vals) >= DETECT_THRESHOLD:
                return label, ora_fmt, str(slice_len)

        return None

    def detect_date_column(
        self, table_name: str, columns: List[Dict]
    ) -> Tuple[Optional[str], str, Optional[str], Optional[str]]:
        date_typed = [c for c in columns if c["type"].upper() in DATE_TYPES]
        if date_typed:
            best   = max(date_typed, key=lambda c: (_name_weight(c["name"]), -c["id"]))
            method = "DATE/TS_패턴" if _name_weight(best["name"]) > 0 else "DATE/TS_기본"
            return best["name"], method, None, None

        char_cols = [
            c for c in columns
            if c["type"].upper() in ("VARCHAR2", "VARCHAR", "CHAR", "NVARCHAR2")
            and 6 <= c["length"] <= 30
        ]
        char_cols_sorted = sorted(char_cols, key=lambda c: (-_name_weight(c["name"]), c["id"]))

        for col in char_cols_sorted:
            result = self._detect_varchar_date_pattern(table_name, col["name"])
            if result:
                label, ora_fmt, slice_len = result
                method = "VARCHAR_패턴" if _name_weight(col["name"]) > 0 else "VARCHAR_감지"
                return col["name"], method, ora_fmt, slice_len

        return None, "NONE", None, None

    def _build_date_expr(
        self, col_name: str, ora_fmt: Optional[str], slice_len: Optional[str]
    ) -> Tuple[str, str]:
        if ora_fmt is None:
            min_expr   = f"TO_CHAR(MIN({col_name}), 'YYYY-MM-DD')"
            where_expr = f"{col_name} >= ADD_MONTHS(SYSDATE, -{RECENT_MONTHS})"
        else:
            from dateutil.relativedelta import relativedelta
            sl        = slice_len or "8"
            col_expr  = f"TO_DATE(SUBSTR({col_name}, 1, {sl}), '{ora_fmt}')"
            min_expr  = f"TO_CHAR(MIN({col_expr}), 'YYYY-MM-DD')"
            threshold = (datetime.now() - relativedelta(months=RECENT_MONTHS))
            thr_str   = threshold.strftime("%Y%m%d%H%M%S")[:int(sl)]
            where_expr = f"SUBSTR({col_name}, 1, {sl}) >= '{thr_str}'"
        return min_expr, where_expr

    def analyze_table(self, table_info: Dict) -> Dict:
        tname   = table_info["table_name"]
        comment = table_info["comment"]
        col_cnt = table_info["col_count"]
        seg_b   = table_info["seg_bytes"]

        result = {
            "테이블명":            tname,
            "테이블설명":          comment,
            "사용여부":            "?",
            "데이터보관최초일자":   "-",
            "1개월평균데이터생성건수": 0,
            "칼럼수(개)":          col_cnt,
            "전체테이블용량":      fmt_gb(seg_b),
            "기준컬럼":            "-",
            "탐지방법":            "-",
            "비고":                "",
        }

        try:
            columns  = self.get_columns(tname)
            col_name, method, ora_fmt, slice_len = self.detect_date_column(tname, columns)

            result["기준컬럼"] = col_name or "-"
            result["탐지방법"] = method

            if col_name:
                min_expr, where_expr = self._build_date_expr(col_name, ora_fmt, slice_len)
                sql = f"""
                    SELECT {min_expr},
                           COUNT(*),
                           COUNT(CASE WHEN {where_expr} THEN 1 END)
                    FROM   {tname}
                """
                row = self._execute_one(sql)
                if row:
                    earliest, total, recent = row
                    result["데이터보관최초일자"]       = fmt_date(earliest)
                    result["사용여부"]                = "Y" if (recent or 0) > 0 else "N"
                    if (recent or 0) > 0:
                        result["1개월평균데이터생성건수"] = round(recent / RECENT_MONTHS)
                    elif total and earliest and earliest != "-":
                        try:
                            dt_e   = datetime.strptime(fmt_date(earliest), "%Y-%m-%d")
                            months = max((datetime.now() - dt_e).days / 30.44, 1)
                            result["1개월평균데이터생성건수"] = round(total / months)
                        except Exception:
                            result["1개월평균데이터생성건수"] = total or 0
            else:
                row   = self._execute_one(f"SELECT COUNT(*) FROM {tname}")
                total = row[0] if row else 0
                result["사용여부"] = "?" if total > 0 else "N"
                result["비고"]     = "날짜 컬럼 없음 — 사용여부 판단 불가"

        except Exception as e:
            result["비고"] = f"ERROR: {str(e)[:100]}"

        return result

    def run(self) -> List[Dict]:
        """테이블별 분석 결과를 List[Dict] 로 반환"""
        print(f"\n[{self.label}] 분석 시작")
        table_list = self.get_table_list()
        total      = len(table_list)
        print(f"  테이블 수: {total}개")

        results = []
        for i, tinfo in enumerate(table_list, 1):
            tname = tinfo["table_name"]
            print(f"  [{i:4d}/{total}] {tname}", end="", flush=True)
            row = self.analyze_table(tinfo)
            results.append(row)
            print(f"  →  {row['사용여부']}  {row['데이터보관최초일자']}")

        print(f"[{self.label}] 완료\n")
        return results


# =============================================================================
# DataFrame 구성
# =============================================================================
# 출력 컬럼 순서 (소재는 combined 파일에만 추가)
_OUT_COLS = [
    "번호", "시스템명", "테이블명", "테이블설명", "사용여부",
    "데이터보관최초일자", "1개월평균데이터생성건수", "칼럼수(개)", "전체테이블용량",
    "기준컬럼", "탐지방법", "비고",
]
_OUT_COLS_COMBINED = [
    "번호", "시스템명", "소재", "테이블명", "테이블설명", "사용여부",
    "데이터보관최초일자", "1개월평균데이터생성건수", "칼럼수(개)", "전체테이블용량",
    "기준컬럼", "탐지방법", "비고",
]


def build_prod_df(prod_results: List[Dict]) -> pd.DataFrame:
    rows = []
    for i, r in enumerate(prod_results, 1):
        row = {"번호": i, "시스템명": SYSTEM_NAME}
        row.update(r)
        rows.append(row)
    df = pd.DataFrame(rows)
    return df[[c for c in _OUT_COLS if c in df.columns]]


def build_combined_df(prod_results: List[Dict], arch_results: List[Dict]) -> pd.DataFrame:
    """
    운영 + 아카이브 합산.
    - 운영에 있는 테이블 → 운영 데이터 그대로 사용
    - 양쪽에 있는 테이블 → 운영 데이터 사용 (소재=양쪽)
    - 아카이브에만 있는 테이블 → 아카이브 데이터 + 사용여부 강제 N (소재=아카이브만)
    """
    prod_names = {r["테이블명"] for r in prod_results}
    arch_names = {r["테이블명"] for r in arch_results}

    combined: List[Tuple[str, Dict]] = []  # (소재, result_dict)

    for r in prod_results:
        source = "양쪽" if r["테이블명"] in arch_names else "운영"
        combined.append((source, r))

    for r in arch_results:
        if r["테이블명"] not in prod_names:
            r_copy = dict(r)
            r_copy["사용여부"] = "N"   # 아카이브에만 있음 = 운영에서 사용 안 함
            combined.append(("아카이브만", r_copy))

    combined.sort(key=lambda x: x[1]["테이블명"])

    rows = []
    for i, (source, r) in enumerate(combined, 1):
        row = {"번호": i, "시스템명": SYSTEM_NAME, "소재": source}
        row.update(r)
        rows.append(row)

    df = pd.DataFrame(rows)
    return df[[c for c in _OUT_COLS_COMBINED if c in df.columns]]


# =============================================================================
# Excel 출력
# =============================================================================
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
Y_FILL      = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
N_FILL      = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
Q_FILL      = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)


def _apply_sheet_style(ws, use_col_name: str = "사용여부"):
    col_idx_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    use_col_idx = col_idx_map.get(use_col_name)

    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER

    for row_idx in range(2, ws.max_row + 1):
        for cell in ws[row_idx]:
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
        if use_col_idx:
            use_cell = ws.cell(row=row_idx, column=use_col_idx)
            val = str(use_cell.value or "")
            if "Y" in val:
                use_cell.fill = Y_FILL
            elif "N" in val:
                use_cell.fill = N_FILL
            elif "?" in val:
                use_cell.fill = Q_FILL

    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 40)

    ws.freeze_panes = "B2"


def save_excel(df: pd.DataFrame, filepath: str, sheet_name: str):
    print(f"  저장 중: {filepath}")
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(filepath)
    _apply_sheet_style(wb[sheet_name])
    wb.save(filepath)
    print(f"  완료: {filepath}  ({len(df)}개 테이블)")


# =============================================================================
# 메인
# =============================================================================
def main():
    try:
        from dateutil.relativedelta import relativedelta  # noqa: F401
    except ImportError:
        print("[ERROR] python-dateutil 패키지가 없습니다: pip install python-dateutil")
        sys.exit(1)

    try:
        oracledb.init_oracle_client(lib_dir=ORACLE_CLIENT_LIB)
        print(f"[Oracle Client] Thick 모드 초기화: {ORACLE_CLIENT_LIB}")
    except Exception as e:
        print(f"[ERROR] Oracle Client 초기화 실패: {e}")
        sys.exit(1)

    print("=" * 60)
    print("  Oracle 테이블 현황 조사")
    print(f"  사용여부 기준: 최근 {RECENT_MONTHS}개월 내 INSERT")
    print(f"  출력 파일: {OUTPUT_PROD}")
    print(f"             {OUTPUT_COMBINED}")
    print("=" * 60)

    # ── 운영DB 분석 ──
    prod_insp = OracleInspector("prod")
    try:
        prod_insp.connect()
        prod_results = prod_insp.run()
    finally:
        prod_insp.close()

    # ── 아카이브DB 분석 ──
    arch_insp = OracleInspector("archive")
    try:
        arch_insp.connect()
        arch_results = arch_insp.run()
    finally:
        arch_insp.close()

    # ── 파일 저장 ──
    print("\n[파일 저장]")

    df_prod = build_prod_df(prod_results)
    save_excel(df_prod, OUTPUT_PROD, sheet_name="운영DB")

    df_combined = build_combined_df(prod_results, arch_results)
    save_excel(df_combined, OUTPUT_COMBINED, sheet_name="운영+아카이브")

    # ── 요약 ──
    print("\n[요약]")
    for label, df in [("운영DB (_prod)", df_prod), ("합산 (_prod+arch)", df_combined)]:
        total = len(df)
        y = (df["사용여부"] == "Y").sum()
        n = (df["사용여부"] == "N").sum()
        q = total - y - n
        print(f"  {label}: 전체 {total}  /  Y={y}  N={n}  ?={q}")


if __name__ == "__main__":
    main()
